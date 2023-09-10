from sys import argv
from datetime import datetime
from openpyxl import load_workbook

path = "C:\\Users\\arsht\\OneDrive\\Desktop\\Resumes\\application_tracker.xlsx"

wb = load_workbook(path)
ws = wb[wb.sheetnames[0]]

today = datetime.now().strftime("%d-%m-%y")

# print("Hi there it is " + date)

n = len(argv)

if (n != 4):
    print("Usage: ./add.py operation(a/u) company position/update")
    print(str(n) + "arguemts passed")
    exit(1)

row = ws['H1'].value + 2
rowstr = str(row)

if (argv[1] == "a"):
    ws['H1'] = row - 1
    srno = 'A' + rowstr
    comp = 'B' + rowstr
    pos  = 'C' + rowstr
    date = 'D' + rowstr
    stat = 'E' + rowstr
    ws[srno] = row - 1
    ws[comp] = argv[2]
    ws[pos]  = argv[3]
    ws[date] = today
    ws[stat] = 'Applied'
    print('Added to the sheet')
elif (argv[1] == "u"):
    print("Updating applciation")
    updated = False
    for i in range(row):
        if (ws['B' + str(i+2)].value == argv[2]):
            ws['D' + str(i+1)] = today
            ws['E' + str(i+2)] = argv[3]
            updated = True
    if (updated):
        print("Updated")
    else:
        print("No entry with company " + argv[2] + " was found")
else:
    print("Incorrect argument passed")

wb.save(filename=path)
